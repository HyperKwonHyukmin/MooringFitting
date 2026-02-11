import os

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


class createReport:
  def __init__(self, element_cls, property_cls, fig_file_dict, MF_dict, winchLC, op2Results, LC_Summary):
    self.element_cls = element_cls
    self.property_cls = property_cls
    self.fig_file_dict = fig_file_dict
    self.MF_dict = MF_dict
    self.winchLC = winchLC
    self.op2Results = op2Results
    self.LC_Summary = LC_Summary

    # 샘플 엑셀 위치 지정
    sample_folder = os.path.dirname(os.path.dirname(os.getcwd()))
    Sample_Excel = os.path.join(sample_folder,"Report_Sample.xlsx" )
    # Sample_Excel = r'C:\Coding\Python\Projects\Mooring_deck\Report_Sample.xlsx'
    self.wb = load_workbook(Sample_Excel)

    # 테두리 스타일 설정 (전체 테두리 적용)
    self.thin_border = Border(
      left=Side(style='thin'),
      right=Side(style='thin'),
      top=Side(style='thin'),
      bottom=Side(style='thin')
    )

    self.highlight_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # 연한 하늘색
    self.center_alignment = Alignment(horizontal="center", vertical="center")  # 중앙 정렬 스타일 생성


  def Run(self, saveFile):
    self.ImportFigure()
    self.ImportPropertyInfo()
    self.ImportLoadMF()
    self.ImportLoadWinch()
    self.ImportLoadCaseSummary()
    self.ImportStressResults()
    self.wb.save(saveFile)


  def ImportFigure(self):
    ws = self.wb['Model']

    MF_fig_start_row = 42
    for file in self.fig_file_dict:
      ws = self.wb['Model']
      img = Image(self.fig_file_dict[file])
      original_width = img.width
      original_height = img.height
      new_width = 1000
      new_height = int(original_height * (new_width / original_width))
      img.width = new_width
      img.height = new_height
      if file == 'full_fig_file':
        ws.cell(row=4, column=1, value="Full Model")
        ws.add_image(img, 'B5')
      else:
        ws.cell(row=MF_fig_start_row - 1, column=1, value=file)
        cellname = 'B' + str(MF_fig_start_row)
        ws.add_image(img, cellname)
        MF_fig_start_row += 38


  def ImportPropertyInfo(self):
    ws = self.wb['Element_Info']
    row = 13

    for i in self.element_cls:
      propertyID = i[1]['property']
      Type = i[1]['type']
      Bt = self.property_cls[propertyID]['rest']['pBeamCreator_dict']['Bt']
      Tt = self.property_cls[propertyID]['rest']['pBeamCreator_dict']['Tt']
      Hw = self.property_cls[propertyID]['rest']['pBeamCreator_dict']['Hw']
      Tw = self.property_cls[propertyID]['rest']['pBeamCreator_dict']['Tw']
      Bb = self.property_cls[propertyID]['rest']['pBeamCreator_dict']['Bb']
      Tb = self.property_cls[propertyID]['rest']['pBeamCreator_dict']['Tb']

      ws.cell(row=row, column=1, value=i[0])
      ws.cell(row=row, column=1).border = self.thin_border
      ws.cell(row=row, column=1).alignment = self.center_alignment
      ws.cell(row=row, column=2, value=Type)
      ws.cell(row=row, column=2).border = self.thin_border
      ws.cell(row=row, column=2).alignment = self.center_alignment
      ws.cell(row=row, column=3, value=Bt)
      ws.cell(row=row, column=3).border = self.thin_border
      ws.cell(row=row, column=3).alignment = self.center_alignment
      ws.cell(row=row, column=4, value=Tt)
      ws.cell(row=row, column=4).border = self.thin_border
      ws.cell(row=row, column=4).alignment = self.center_alignment
      ws.cell(row=row, column=5, value=Hw)
      ws.cell(row=row, column=5).border = self.thin_border
      ws.cell(row=row, column=5).alignment = self.center_alignment
      ws.cell(row=row, column=6, value=Tw)
      ws.cell(row=row, column=6).border = self.thin_border
      ws.cell(row=row, column=6).alignment = self.center_alignment
      ws.cell(row=row, column=7, value=Bb)
      ws.cell(row=row, column=7).border = self.thin_border
      ws.cell(row=row, column=7).alignment = self.center_alignment
      ws.cell(row=row, column=8, value=Tb)
      ws.cell(row=row, column=8).border = self.thin_border
      ws.cell(row=row, column=8).alignment = self.center_alignment
      row += 1


  def ImportLoadMF(self):
    ws = self.wb['Load_MF']

    row = 4  # 4행부터 시작
    for ID in self.MF_dict:
      ws.cell(row=row, column=1, value=self.MF_dict[ID]['Name'])
      ws.cell(row=row, column=1).border = self.thin_border
      ws.cell(row=row, column=1).alignment = self.center_alignment
      ws.cell(row=row, column=2, value=self.MF_dict[ID]['Type'])
      ws.cell(row=row, column=2).border = self.thin_border
      ws.cell(row=row, column=2).alignment = self.center_alignment
      ws.cell(row=row, column=3, value=self.MF_dict[ID]['X'])
      ws.cell(row=row, column=3).border = self.thin_border
      ws.cell(row=row, column=3).alignment = self.center_alignment
      ws.cell(row=row, column=4, value=self.MF_dict[ID]['Y'])
      ws.cell(row=row, column=4).border = self.thin_border
      ws.cell(row=row, column=4).alignment = self.center_alignment
      ws.cell(row=row, column=5, value=self.MF_dict[ID]['Z'])
      ws.cell(row=row, column=5).border = self.thin_border
      ws.cell(row=row, column=5).alignment = self.center_alignment
      ws.cell(row=row, column=6, value=self.MF_dict[ID]['SWL'])
      ws.cell(row=row, column=6).border = self.thin_border
      ws.cell(row=row, column=6).alignment = self.center_alignment
      ws.cell(row=row, column=7, value=self.MF_dict[ID]['a'])
      ws.cell(row=row, column=7).border = self.thin_border
      ws.cell(row=row, column=7).alignment = self.center_alignment
      ws.cell(row=row, column=8, value=self.MF_dict[ID]['b'])
      ws.cell(row=row, column=8).border = self.thin_border
      ws.cell(row=row, column=8).alignment = self.center_alignment
      ws.cell(row=row, column=9, value=self.MF_dict[ID]['c'])
      ws.cell(row=row, column=9).border = self.thin_border
      ws.cell(row=row, column=9).alignment = self.center_alignment
      ws.cell(row=row, column=10, value=self.MF_dict[ID]['MF_degree'])
      ws.cell(row=row, column=10).border = self.thin_border
      ws.cell(row=row, column=10).alignment = self.center_alignment
      ws.cell(row=row, column=11, value=self.MF_dict[ID]['Force_X'])
      ws.cell(row=row, column=11).border = self.thin_border
      ws.cell(row=row, column=11).alignment = self.center_alignment
      ws.cell(row=row, column=11).fill = self.highlight_fill
      ws.cell(row=row, column=12, value=self.MF_dict[ID]['Force_Y'])
      ws.cell(row=row, column=12).border = self.thin_border
      ws.cell(row=row, column=12).alignment = self.center_alignment
      ws.cell(row=row, column=12).fill = self.highlight_fill
      row += 1


  def ImportLoadWinch(self):
    ws = self.wb['Load_Winch']

    F_list = []
    # 일단 첫번째 열에 F1, F2 ... 입력하기
    row = 4  # 4행부터 시작
    for LC in self.winchLC['Calculated_force']:
      for F in self.winchLC['Calculated_force'][LC]:
        ws.cell(row=row, column=1, value=F)
        ws.cell(row=row, column=1).border = self.thin_border
        ws.cell(row=row, column=1).alignment = self.center_alignment
        row += 1
        F_list.append(F)
      break

    # wichLC 입력 시작
    column = 2
    for LC in self.winchLC['Calculated_force']:
      row = 4
      for F in F_list:
        ws.cell(row=row, column=column, value=self.winchLC['Calculated_force'][LC][F]['forceX'])
        ws.cell(row=row, column=column).border = self.thin_border
        ws.cell(row=row, column=column).alignment = self.center_alignment
        ws.cell(row=row, column=column + 1, value=self.winchLC['Calculated_force'][LC][F]['forceY'])
        ws.cell(row=row, column=column + 1).border = self.thin_border
        ws.cell(row=row, column=column + 1).alignment = self.center_alignment
        ws.cell(row=row, column=column + 2, value=self.winchLC['Calculated_force'][LC][F]['forceZ'])
        ws.cell(row=row, column=column + 2).border = self.thin_border
        ws.cell(row=row, column=column + 2).alignment = self.center_alignment
        row += 1
      column += 3


  def ImportLoadCaseSummary(self):
    ws = self.wb['LoadCase_Summary']
    row = 2
    for i, s in enumerate(self.LC_Summary):
      LC = 'Load' + str(i + 1)
      ws.cell(row=row, column=1, value=LC)
      ws.cell(row=row, column=1).border = self.thin_border
      ws.cell(row=row, column=1).alignment = self.center_alignment
      ws.cell(row=row, column=2, value=s)
      ws.cell(row=row, column=2).border = self.thin_border
      ws.cell(row=row, column=2).alignment = self.center_alignment
      row += 1


  def ImportStressResults(self):
    LC_sheet_name = 'Results_LC1'
    original_ws = self.wb[LC_sheet_name]
    LC_count = len(self.op2Results.ELForce_results)

    for LC in range(2, LC_count + 1):  # 일단 결과 Load Case 만큼 Sheet 복사해두기
      new_LC_sheet_name = LC_sheet_name[:-1] + str(LC)
      copied_ws = self.wb.copy_worksheet(original_ws)
      copied_ws.title = new_LC_sheet_name
      LC_text = 'LC ' + str(LC - 1 + 1)
      copied_ws.cell(row=1, column=1, value=LC_text)

    start_row, start_col = 10, 1

    # 구조해석 결과 입력 시작
    for i in range(LC_count):
      ws = self.wb[f'Results_LC{i+1}']

      # 특정 열 선택 및 순서 재배치
      selected_columns = ['Element', 'Node','Bending_1', 'Bending_2', 'Shear_1', 'Shear_2',
                                           'Axial', 'Torque', 'Nx', 'Mx', 'Qy', 'Qz', 'My', 'Mz']  # 원하는 열 순서
      df_selected = self.op2Results.ELForce_results[i][selected_columns]  # 선택한 열만 포함
      for r_idx, row in enumerate(dataframe_to_rows(df_selected, index=False, header=False), start=start_row):
        for c_idx, value in enumerate(row, start=start_col):
          ws.cell(row=r_idx, column=c_idx, value=value)

